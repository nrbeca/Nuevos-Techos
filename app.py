import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Distribución Nuevo Techo AC01", page_icon="📊", layout="wide")

# ----------------------------------------------------------------------------
# Columnas de la hoja "Base Estimación"
# ----------------------------------------------------------------------------
BASE_COLS = [
    "UR", "Nombre UR", "Fi", "Fn", "SF", "AI", "PP", "PE", "Nombre Partida",
    "Capítulo", "TG", "FF", "EF", "CC", "Monto Anual",
]

HEADER_FILL = PatternFill(start_color="941100", end_color="941100", fill_type="solid")
HEADER_FONT = Font(name="Montserrat", size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Montserrat", size=10, bold=True)
MONEY_FMT = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
PCT_FMT = "0.0%"

# Partidas específicas que se excluyen siempre del cálculo (todas las UR)
PARTIDAS_EXCLUIDAS = {39801, 39401}


@st.cache_data(show_spinner=False)
def load_workbook_data(file_bytes: bytes):
    """Lee la Base Estimación y los Nuevos Techos del archivo cargado."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))

    base = pd.read_excel(xls, sheet_name="Base Estimación", header=1)
    base = base[[c for c in BASE_COLS if c in base.columns]].copy()
    base = base.dropna(subset=["UR"])
    base["UR"] = base["UR"].astype(int)

    # Excluir partidas específicas en todas las UR (39801 y 39401)
    base["PE"] = pd.to_numeric(base["PE"], errors="coerce")
    base = base[~base["PE"].isin(PARTIDAS_EXCLUIDAS)].copy()

    techos_raw = pd.read_excel(xls, sheet_name="Nuevos Techos", header=3)
    techos_raw = techos_raw.dropna(subset=["UR"])
    techos_raw["UR"] = techos_raw["UR"].astype(int)
    techos = dict(zip(techos_raw["UR"], techos_raw["Nuevo Techo"]))

    return base, techos


def _write_ur_sheet(ws, df_ur: pd.DataFrame, ur: int, nombre_ur: str, nuevo_techo: float):
    """Escribe en una hoja (ya creada) el detalle de una UR, con fórmulas vivas."""
    n_rows = len(df_ur)
    first_data_row = 6
    last_data_row = first_data_row + n_rows - 1
    total_row = last_data_row + 1

    # --- Encabezado con el nuevo techo ---
    ws["Q3"] = "Nuevo Techo"
    ws["Q3"].font = LABEL_FONT
    ws["R3"] = float(nuevo_techo)
    ws["R3"].number_format = MONEY_FMT

    ws["A1"] = f"UR {ur} - {nombre_ur}"
    ws["A1"].font = Font(name="Montserrat", size=12, bold=True)

    # --- Encabezados de columnas (fila 5) ---
    header_row = 5
    headers = BASE_COLS + ["Porcentaje", None, "Nuevo Techo", "Nuevo techo redondeado"]
    for idx, h in enumerate(headers, start=1):
        if h is None:
            continue
        cell = ws.cell(row=header_row, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_letters = {name: get_column_letter(i) for i, name in enumerate(BASE_COLS, start=1)}
    P_COL = get_column_letter(len(BASE_COLS) + 1)  # Porcentaje
    R_COL = get_column_letter(len(BASE_COLS) + 3)  # Nuevo Techo
    S_COL = get_column_letter(len(BASE_COLS) + 4)  # Nuevo techo redondeado
    O_COL = col_letters["Monto Anual"]

    # --- Filas de datos con fórmulas ---
    for i, (_, row) in enumerate(df_ur.iterrows()):
        r = first_data_row + i
        for name in BASE_COLS:
            c = ws.cell(row=r, column=BASE_COLS.index(name) + 1, value=row[name])
            if name == "Monto Anual":
                c.number_format = MONEY_FMT

        ws[f"{P_COL}{r}"] = f"=+{O_COL}{r}/${O_COL}${total_row}"
        ws[f"{P_COL}{r}"].number_format = PCT_FMT

        ws[f"{R_COL}{r}"] = f"=+$R$3*{P_COL}{r}"
        ws[f"{R_COL}{r}"].number_format = MONEY_FMT

        ws[f"{S_COL}{r}"] = f"=+ROUND({R_COL}{r},0)"
        ws[f"{S_COL}{r}"].number_format = MONEY_FMT

    # --- Fila de totales ---
    n_col_letter = col_letters["CC"]
    ws[f"{n_col_letter}{total_row}"] = "TOTAL"
    ws[f"{n_col_letter}{total_row}"].font = LABEL_FONT
    ws[f"{O_COL}{total_row}"] = f"=SUM({O_COL}{first_data_row}:{O_COL}{last_data_row})"
    ws[f"{O_COL}{total_row}"].number_format = MONEY_FMT
    ws[f"{O_COL}{total_row}"].font = LABEL_FONT

    ws[f"{S_COL}{total_row}"] = f"=SUM({S_COL}{first_data_row}:{S_COL}{last_data_row})"
    ws[f"{S_COL}{total_row}"].number_format = MONEY_FMT
    ws[f"{S_COL}{total_row}"].font = LABEL_FONT

    # --- Ancho de columnas ---
    widths = [8, 26, 5, 5, 5, 6, 7, 8, 40, 9, 5, 5, 5, 14, 15, 11, 3, 15, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{first_data_row}"


def _sheet_name_for_ur(ur: int, used: set) -> str:
    """Genera un nombre de hoja único y válido (máx. 31 caracteres) para la UR."""
    name = f"UR {ur}"
    base_name = name
    n = 1
    while name in used:
        n += 1
        name = f"{base_name} ({n})"
    used.add(name)
    return name[:31]


def build_excel(df_ur: pd.DataFrame, ur: int, nombre_ur: str, nuevo_techo: float) -> bytes:
    """Genera el Excel descargable (una sola UR) con fórmulas vivas."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"UR {ur}"[:31]
    _write_ur_sheet(ws, df_ur, ur, nombre_ur, nuevo_techo)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_excel_all(base_df: pd.DataFrame, techos: dict, ur_nombres: pd.DataFrame) -> bytes:
    """Genera un solo Excel con una hoja por cada UR que tenga Nuevo Techo asignado."""
    wb = Workbook()
    wb.remove(wb.active)  # se irán agregando las hojas de cada UR

    used_names = set()
    urs_ordenadas = sorted(techos.keys())
    for ur in urs_ordenadas:
        df_ur = base_df[base_df["UR"] == ur].reset_index(drop=True)
        if df_ur.empty:
            continue  # UR con techo pero sin partidas en la Base Estimación
        nombre_ur = ur_nombres.loc[ur_nombres["UR"] == ur, "Nombre UR"]
        nombre_ur = nombre_ur.iloc[0] if len(nombre_ur) else ""
        nuevo_techo = techos[ur]

        sheet_name = _sheet_name_for_ur(ur, used_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_ur_sheet(ws, df_ur, ur, nombre_ur, nuevo_techo)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_excel_consolidado(base_df: pd.DataFrame, techos: dict, ur_nombres: pd.DataFrame) -> bytes:
    """Genera un Excel de una sola hoja con todas las UR corridas en orden,
    solo con valores (sin fórmulas)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    headers = BASE_COLS + ["Porcentaje", "Nuevo Techo", "Nuevo techo redondeado"]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    PCT_IDX = len(BASE_COLS) + 1
    NT_IDX = len(BASE_COLS) + 2
    NTR_IDX = len(BASE_COLS) + 3

    r = 2
    urs_ordenadas = sorted(techos.keys())
    for ur in urs_ordenadas:
        df_ur = base_df[base_df["UR"] == ur].reset_index(drop=True)
        if df_ur.empty:
            continue

        nuevo_techo = techos[ur]
        total_monto = df_ur["Monto Anual"].sum()

        for _, row in df_ur.iterrows():
            for idx, name in enumerate(BASE_COLS, start=1):
                c = ws.cell(row=r, column=idx, value=row[name])
                if name == "Monto Anual":
                    c.number_format = MONEY_FMT

            porcentaje = (row["Monto Anual"] / total_monto) if total_monto else 0
            nuevo_techo_partida = nuevo_techo * porcentaje
            nuevo_techo_redondeado = round(nuevo_techo_partida)

            c = ws.cell(row=r, column=PCT_IDX, value=porcentaje)
            c.number_format = PCT_FMT

            c = ws.cell(row=r, column=NT_IDX, value=nuevo_techo_partida)
            c.number_format = MONEY_FMT

            c = ws.cell(row=r, column=NTR_IDX, value=nuevo_techo_redondeado)
            c.number_format = MONEY_FMT

            r += 1

    # --- Ancho de columnas ---
    widths = [8, 26, 5, 5, 5, 6, 7, 8, 40, 9, 5, 5, 5, 14, 15, 11, 15, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ----------------------------------------------------------------------------
# Interfaz
# ----------------------------------------------------------------------------
def main():
    st.title("Distribución del Nuevo Techo AC01 por Unidad Responsable")
    st.caption(
        "Carga la base (hojas 'Base Estimación' y 'Nuevos Techos'), elige la UR y descarga "

    )

    uploaded = st.file_uploader("Sube el archivo base (.xlsx)", type=["xlsx"])

    if uploaded is None:
        st.info("Sube el archivo con las hojas 'Base Estimación' y 'Nuevos Techos' para comenzar.")
        st.stop()

    try:
        base_df, techos = load_workbook_data(uploaded.getvalue())
    except Exception as e:
        st.error(f"No se pudo leer el archivo. Verifica que tenga las hojas 'Base Estimación' y 'Nuevos Techos'. Detalle: {e}")
        st.stop()

    # Mapa UR -> Nombre UR
    ur_nombres = (
        base_df[["UR", "Nombre UR"]]
        .drop_duplicates()
        .sort_values("UR")
        .assign(label=lambda d: d["UR"].astype(str) + " - " + d["Nombre UR"].astype(str))
    )

    urs_sin_techo = sorted(set(ur_nombres["UR"]) - set(techos.keys()))

    st.sidebar.header("Selección")
    seleccion = st.sidebar.selectbox("Unidad Responsable (UR)", ur_nombres["label"].tolist())
    ur_sel = int(seleccion.split(" - ")[0])
    nombre_ur_sel = ur_nombres.loc[ur_nombres["UR"] == ur_sel, "Nombre UR"].iloc[0]

    if ur_sel in urs_sin_techo:
        st.warning(
            f"La UR {ur_sel} - {nombre_ur_sel} no tiene un Nuevo Techo asignado en la hoja "
            "'Nuevos Techos'. Agrega el valor correspondiente antes de generar el reporte."
        )
        st.stop()

    nuevo_techo_ur = techos[ur_sel]
    df_ur = base_df[base_df["UR"] == ur_sel].reset_index(drop=True)

    total_monto = df_ur["Monto Anual"].sum()
    df_ur["Porcentaje"] = df_ur["Monto Anual"] / total_monto
    df_ur["Nuevo Techo"] = nuevo_techo_ur * df_ur["Porcentaje"]
    df_ur["Nuevo techo redondeado"] = df_ur["Nuevo Techo"].round(0)

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Monto Anual (UR)", f"$ {total_monto:,.2f}")
    c2.metric("Nuevo Techo asignado", f"$ {nuevo_techo_ur:,.2f}")
    c3.metric("Nuevo techo redondeado (suma)", f"$ {df_ur['Nuevo techo redondeado'].sum():,.0f}")

    st.subheader(f"UR {ur_sel} - {nombre_ur_sel}")
    st.dataframe(
        df_ur.style.format(
            {
                "Monto Anual": "{:,.2f}",
                "Porcentaje": "{:.1%}",
                "Nuevo Techo": "{:,.2f}",
                "Nuevo techo redondeado": "{:,.0f}",
            }
        ),
        use_container_width=True,
        height=450,
    )

    excel_bytes = build_excel(df_ur, ur_sel, nombre_ur_sel, nuevo_techo_ur)
    st.download_button(
        label=" Descargar Excel con fórmulas (esta UR)",
        data=excel_bytes,
        file_name=f"AC01_UR_{ur_sel}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()
    st.subheader("Todas las UR en un solo archivo")
    urs_con_techo = sorted(set(techos.keys()) & set(base_df["UR"]))
    st.write(
        f"Se generará un Excel con {len(urs_con_techo)} hojas, una por cada UR que tiene "

    )
    if st.button("Generar Excel con todas las UR"):
        with st.spinner("Generando el archivo con todas las UR..."):
            all_bytes = build_excel_all(base_df, techos, ur_nombres)
        st.download_button(
            label=" Descargar Excel — todas las UR",
            data=all_bytes,
            file_name=f"AC01_todas_las_UR_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.divider()
    st.subheader("Consolidado en una sola hoja")
    st.write(
        "Genera un único Excel con todas las partidas de todas las UR corridas en orden "

    )
    if st.button("Generar Excel consolidado"):
        with st.spinner("Generando el archivo consolidado..."):
            consolidado_bytes = build_excel_consolidado(base_df, techos, ur_nombres)
        st.download_button(
            label=" Descargar Excel consolidado",
            data=consolidado_bytes,
            file_name=f"AC01_consolidado_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if urs_sin_techo:
        with st.sidebar.expander(" URs sin Nuevo Techo"):
            st.write(", ".join(str(u) for u in urs_sin_techo))


if __name__ == "__main__":
    main()
